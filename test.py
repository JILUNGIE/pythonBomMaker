import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import *

COLUMN_ITEM_POS = 1
COLUMN_PART_POS = 2
COLUMN_QUANTITY_POS = 3
COLUMN_REFERENCE_POS = 4
# COLUMN_REMARK_POS = 5
# COLUMN_NOTE_POS = 6

COLUMN_ITEM_VALUE_INDEX = 0
COLUMN_PART_VALUE_INDEX = 3
COLUMN_QUANTITY_VALUE_INDEX = 1
COLUMN_REFERENCE_VALUE_INDEX = 2

COLUMN_POS = [COLUMN_ITEM_POS, COLUMN_PART_POS, COLUMN_QUANTITY_POS, COLUMN_REFERENCE_POS] # BOM 목록 순서
COLUMN_VALUE_INDEX = [COLUMN_ITEM_VALUE_INDEX, COLUMN_PART_VALUE_INDEX, COLUMN_QUANTITY_VALUE_INDEX, COLUMN_REFERENCE_VALUE_INDEX] # Excel 목록 순서
ROW1_VALUE = ["Item","Part", "Quantity", "Reference", "Remark", "비고"]

month = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06',
         'July': '07', 'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12'}

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


def fileExplorer():  # https://stackoverflow.com/questions/9319317/quick-and-easy-file-dialog-in-python
    filetypes = (('bom files', '*.bom'),)
    root = tk.Tk()
    root.withdraw()
    # https://www.pythontutorial.net/tkinter/tkinter-open-file-dialog/
    files = filedialog.askopenfilename(
        title='Select bom file', filetypes=filetypes)

    if files == '':
        messagebox.showwarning("Warning!!", "Please select a file")
    return files


def fileName(title):
    bomContents = title.split()

    bomTitle = bomContents[0]
    bomYear = bomContents[5]
    bomMonth = month[bomContents[3]]
    bomDate = bomContents[4].replace(',', '')

    bomBirthDay = "%s(%s.%s.%s)" % (bomTitle, bomYear, bomMonth, bomDate)

    return bomBirthDay

def bomFormat():
    ws["A1"] = "Item"
    for i in range(0,6):
        ws.cell(row=1, column=i+1, value=ROW1_VALUE[i])
        ws.cell(row=1, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=i+1).border = thin_border
        ws.cell(row=1, column=i+1).fill = PatternFill(fill_type='solid', fgColor=Color('C6E0B4'))
    ws.column_dimensions["B"].width = 33
    ws.column_dimensions["D"].width = 80

if __name__ == "__main__":
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"  # sheet 의 이름 변경

    file = open(fileExplorer(), 'r')
    if file.readable == False:
        messagebox.showwarning("Warning!!", "Can't read file!")

    filedocument = file.readlines()
    introLine = filedocument[0]
    lineLen = len(filedocument)

    countLine = 15
    row_value = 2

    bomFormat()
    while True:
        if countLine > lineLen:
            break
        docs = filedocument[countLine-1].split()
        for i in range(0,4):
            if i%2==1:
                ws.cell(row=row_value, column=COLUMN_POS[i],value=docs[COLUMN_VALUE_INDEX[i]])
            else:
                ws.cell(row=row_value, column=COLUMN_POS[i],value=int(docs[COLUMN_VALUE_INDEX[i]]))
            ws.cell(row=row_value, column=COLUMN_POS[i]).border = thin_border
        ws.cell(row=row_value, column=1).alignment = Alignment(horizontal='center', vertical='center')
        
        countLine = countLine+1
        row_value = row_value+1

    bomtitle = "%s.xlsx" % (fileName(introLine))
    file.close()
    wb.save(bomtitle)
    wb.close()